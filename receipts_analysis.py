"""
PAC Receipts Analysis
======================
Takes one or more Schedule A receipts CSVs (one per corporation, each
covering multiple election cycles) and summarizes contribution totals by:
  - Election cycle
  - Year
  - Quarter
  - Year-Quarter

Each corporation gets its own set of tabs in a single Excel output.
A summary sheet compares all corporations side by side.

Requirements:
    pip install pandas openpyxl tkinter

Usage:
    python receipts_analysis.py
    (a file picker will open for multi-select)
"""

import sys
from pathlib import Path
import pandas as pd

# ── Config ────────────────────────────────────────────────────────────────────
OUTPUT_FILE    = "pac_receipts_analysis.xlsx"
AMOUNT_COL     = "contribution_receipt_amount"
DATE_COL       = "contribution_receipt_date"
CYCLE_COL      = "two_year_transaction_period"
COMMITTEE_COL  = "committee_name"

# Map FEC two-year period → readable label
CYCLE_LABELS = {
    2022: "2021-2022",
    2024: "2023-2024",
    2026: "2025-2026",
    2020: "2019-2020",
    2018: "2017-2018",
}
# ─────────────────────────────────────────────────────────────────────────────


# ── File picker ───────────────────────────────────────────────────────────────
def pick_input_files():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        paths = filedialog.askopenfilenames(
            title="Select one or more PAC receipts CSVs",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        root.destroy()
        return list(paths)
    except Exception as e:
        print(f"  [info] file picker unavailable ({e}); falling back to typed input")
        typed = input("Path(s) to CSV(s), comma-separated: ").strip()
        return [p.strip().strip('"') for p in typed.split(",") if p.strip()]


# ── Load and parse one CSV ────────────────────────────────────────────────────
def load_csv(path):
    df = pd.read_csv(path, dtype=str, low_memory=False)
    df.columns = [c.strip() for c in df.columns]

    # Amount
    if AMOUNT_COL not in df.columns:
        raise ValueError(f"Missing column '{AMOUNT_COL}' in {path}")
    df[AMOUNT_COL] = pd.to_numeric(df[AMOUNT_COL], errors="coerce")
    df = df.dropna(subset=[AMOUNT_COL])

    # Date → year, quarter
    if DATE_COL in df.columns:
        dates = pd.to_datetime(df[DATE_COL], errors="coerce")
        df["_year"]    = dates.dt.year
        df["_quarter"] = "Q" + dates.dt.quarter.astype(str)
        df["_yq"]      = df["_year"].astype(str) + " " + df["_quarter"]
    else:
        df["_year"] = df["_quarter"] = df["_yq"] = None

    # Election cycle label
    if CYCLE_COL in df.columns:
        df["_cycle_raw"] = pd.to_numeric(df[CYCLE_COL], errors="coerce")
        df["_cycle"] = df["_cycle_raw"].map(
            lambda x: CYCLE_LABELS.get(int(x), str(int(x))) if pd.notna(x) else "Unknown"
        )
    else:
        df["_cycle"] = "Unknown"

    # Corporation name from committee_name column (use filename as fallback)
    if COMMITTEE_COL in df.columns:
        corp_name = df[COMMITTEE_COL].dropna().iloc[0] if not df[COMMITTEE_COL].dropna().empty else Path(path).stem
    else:
        corp_name = Path(path).stem

    return df, corp_name


# ── Summarize one corporation's data ──────────────────────────────────────────
def summarize(df, corp_name):
    amt = AMOUNT_COL

    def agg(group_cols):
        return (
            df.groupby(group_cols)[amt]
            .agg(total_amount="sum", transaction_count="count")
            .reset_index()
        )

    # By election cycle
    by_cycle = agg(["_cycle"])
    by_cycle.rename(columns={"_cycle": "election_cycle"}, inplace=True)
    by_cycle.sort_values("election_cycle", inplace=True)

    # By year
    by_year = df.dropna(subset=["_year"]).copy()
    by_year["_year"] = by_year["_year"].astype(int)
    by_year = (
        by_year.groupby("_year")[amt]
        .agg(total_amount="sum", transaction_count="count")
        .reset_index()
        .rename(columns={"_year": "year"})
        .sort_values("year")
    )

    # By quarter (within each year)
    by_quarter = df.dropna(subset=["_year", "_quarter"]).copy()
    by_quarter["_year"] = by_quarter["_year"].astype(int)
    by_quarter = (
        by_quarter.groupby(["_year", "_quarter"])[amt]
        .agg(total_amount="sum", transaction_count="count")
        .reset_index()
        .rename(columns={"_year": "year", "_quarter": "quarter"})
        .sort_values(["year", "quarter"])
    )

    # By year-quarter (chronological string)
    by_yq = df.dropna(subset=["_yq", "_year", "_quarter"]).copy()
    by_yq["_year"] = by_yq["_year"].astype(int)
    by_yq = (
        by_yq.groupby(["_year", "_quarter", "_yq"])[amt]
        .agg(total_amount="sum", transaction_count="count")
        .reset_index()
        .rename(columns={"_year": "year", "_quarter": "quarter", "_yq": "year_quarter"})
        .sort_values(["year", "quarter"])
    )

    # By cycle + year
    by_cycle_year = df.dropna(subset=["_year"]).copy()
    by_cycle_year["_year"] = by_cycle_year["_year"].astype(int)
    by_cycle_year = (
        by_cycle_year.groupby(["_cycle", "_year"])[amt]
        .agg(total_amount="sum", transaction_count="count")
        .reset_index()
        .rename(columns={"_cycle": "election_cycle", "_year": "year"})
        .sort_values(["election_cycle", "year"])
    )

    # By cycle + quarter
    by_cycle_quarter = df.dropna(subset=["_year", "_quarter"]).copy()
    by_cycle_quarter["_year"] = by_cycle_quarter["_year"].astype(int)
    by_cycle_quarter = (
        by_cycle_quarter.groupby(["_cycle", "_year", "_quarter", "_yq"])[amt]
        .agg(total_amount="sum", transaction_count="count")
        .reset_index()
        .rename(columns={"_cycle": "election_cycle", "_year": "year",
                         "_quarter": "quarter", "_yq": "year_quarter"})
        .sort_values(["election_cycle", "year", "quarter"])
    )

    return {
        "by_cycle":         by_cycle,
        "by_year":          by_year,
        "by_quarter":       by_quarter,
        "by_year_quarter":  by_yq,
        "by_cycle_year":    by_cycle_year,
        "by_cycle_quarter": by_cycle_quarter,
    }


# ── Format dollars in a worksheet ────────────────────────────────────────────
def format_sheet(ws, dollar_cols):
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        header = col[0].value or ""
        col_letter = get_column_letter(col_idx)
        if header in dollar_cols:
            for cell in ws[col_letter][1:]:
                cell.number_format = '$#,##0.00'
        max_len = max((len(str(c.value or "")) for c in ws[col_letter]), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ── Write one corporation to the Excel writer ─────────────────────────────────
def write_corp_sheets(writer, corp_name, summaries):
    # Truncate sheet name prefix to fit Excel's 31-char limit
    prefix = corp_name[:18].strip()

    sheet_map = {
        f"{prefix} — By Cycle":          ("by_cycle",         ["total_amount"]),
        f"{prefix} — By Year":           ("by_year",          ["total_amount"]),
        f"{prefix} — By Quarter":        ("by_quarter",       ["total_amount"]),
        f"{prefix} — By YearQtr":        ("by_year_quarter",  ["total_amount"]),
        f"{prefix} — Cycle+Year":        ("by_cycle_year",    ["total_amount"]),
        f"{prefix} — Cycle+Qtr":         ("by_cycle_quarter", ["total_amount"]),
    }

    for sheet_name, (key, dollar_cols) in sheet_map.items():
        df = summaries[key].copy()
        # Drop internal year/quarter sort cols if present alongside year_quarter
        if "year_quarter" in df.columns and "year" in df.columns and "quarter" in df.columns:
            df.drop(columns=["year", "quarter"], inplace=True, errors="ignore")
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        format_sheet(ws, dollar_cols)


# ── Cross-corporation summary sheet ───────────────────────────────────────────
def write_summary_sheet(writer, all_summaries):
    """One sheet comparing all corporations by cycle."""
    frames = []
    for corp_name, summaries in all_summaries.items():
        df = summaries["by_cycle"].copy()
        df.insert(0, "corporation", corp_name)
        frames.append(df)

    if not frames:
        return

    df_all = pd.concat(frames, ignore_index=True)
    df_all.to_excel(writer, index=False, sheet_name="Summary — All Corps")
    ws = writer.sheets["Summary — All Corps"]
    format_sheet(ws, ["total_amount"])

    # Pivot: corporations as rows, cycles as columns
    pivot = df_all.pivot_table(
        index="corporation", columns="election_cycle",
        values="total_amount", aggfunc="sum", fill_value=0
    ).reset_index()
    pivot.to_excel(writer, index=False, sheet_name="Summary — Pivot")
    ws2 = writer.sheets["Summary — Pivot"]
    format_sheet(ws2, [c for c in pivot.columns if c != "corporation"])


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    input_paths = pick_input_files()
    if not input_paths:
        print("No files selected. Exiting.")
        sys.exit(0)

    all_summaries = {}

    for path in input_paths:
        print(f"Processing: {Path(path).name}")
        try:
            df, corp_name = load_csv(path)
            print(f"  Corporation: {corp_name}")
            print(f"  Rows loaded: {len(df):,}")
            print(f"  Cycles found: {sorted(df['_cycle'].unique())}")
            summaries = summarize(df, corp_name)
            all_summaries[corp_name] = summaries
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

    if not all_summaries:
        print("No files processed successfully. Exiting.")
        sys.exit(1)

    output_path = Path(input_paths[0]).resolve().parent / OUTPUT_FILE
    print(f"\nWriting to {output_path}...")

    # Combine all corporations into one DataFrame per summary type
    combined = {}
    for key in ["by_cycle", "by_year", "by_quarter", "by_year_quarter",
                "by_cycle_year", "by_cycle_quarter"]:
        frames = []
        for corp_name, summaries in all_summaries.items():
            df = summaries[key].copy()
            df.insert(0, "corporation", corp_name)
            frames.append(df)
        combined[key] = pd.concat(frames, ignore_index=True)

    sheet_map = {
        "By Cycle":         ("by_cycle",         ["total_amount"]),
        "By Year":          ("by_year",          ["total_amount"]),
        "By Quarter":       ("by_quarter",       ["total_amount"]),
        "By Year-Quarter":  ("by_year_quarter",  ["total_amount"]),
        "By Cycle + Year":  ("by_cycle_year",    ["total_amount"]),
        "By Cycle + Qtr":   ("by_cycle_quarter", ["total_amount"]),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Pivot summary first
        pivot = combined["by_cycle"].pivot_table(
            index="corporation", columns="election_cycle",
            values="total_amount", aggfunc="sum", fill_value=0
        ).reset_index()
        pivot.to_excel(writer, index=False, sheet_name="Summary Pivot")
        format_sheet(writer.sheets["Summary Pivot"],
                     [c for c in pivot.columns if c != "corporation"])

        # One tab per summary type, all corporations stacked
        for sheet_name, (key, dollar_cols) in sheet_map.items():
            df = combined[key].copy()
            if "year_quarter" in df.columns and "year" in df.columns and "quarter" in df.columns:
                df.drop(columns=["year", "quarter"], inplace=True, errors="ignore")
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_sheet(writer.sheets[sheet_name], dollar_cols)

    print(f"Done! Saved to: {output_path}")
    print(f"\nSheets written:")
    print(f"  Summary Pivot       (corp × cycle pivot)")
    for sheet_name in sheet_map:
        print(f"  {sheet_name}")


if __name__ == "__main__":
    main()