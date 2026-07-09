"""
Disbursement File Classifier — Batch Runner (with Candidate Info)
==================================================================
Classifies every disbursement row as:
  - Authorized Campaign Committee  (entity_type_desc == CAMPAIGN COMMITTEE)
  - Leadership PAC                 (entity_type_desc == POLITICAL ACTION COMMITTEE
                                    AND recipient_committee_id in LPAC master)

Candidate info (name, office, state, district, party, candidate_id)
is joined from the exploded LPAC master via recipient_committee_id.

ALL campaign committee rows are classified as Authorized regardless of
whether the recipient candidate appears in the LPAC master.

Inputs:
  fec_lpac_exploded.xlsx  — exploded master (one row per committee)
  pac_disbursements/      — folder of Schedule B CSVs

Output per CSV:
  Sheets: All Disbursements | Matched Only | Auth Committees |
          Leadership PACs | Democrat | Republican | Unmatched (if any)

Also writes batch_summary.xlsx and batch_log.xlsx.

Requirements: pip install pandas openpyxl
"""

import os, glob, zipfile, io
import pandas as pd
from collections import Counter

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FOLDER      = "pac_disbursements"
OUTPUT_FOLDER     = "disbursements_analyzed_wLPAC"
LPAC_EXPLODED     = "fec_lpac_exploded.xlsx"
TEST_FILE         = ""          # e.g. "raytheon.csv", or "" for all
KEEP_ENTITY_TYPES = {"POLITICAL ACTION COMMITTEE", "CAMPAIGN COMMITTEE"}
# ─────────────────────────────────────────────────────────────────────────────

os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# ── Helpers ───────────────────────────────────────────────────────────────────
def clean(val):
    if pd.isna(val):
        return ""
    return str(val).strip().strip("\r\n").replace("\ufeff", "").upper()


def norm_party(val):
    if pd.isna(val) or str(val).strip() == "":
        return None
    v = str(val).upper()
    if "DEMOCRAT" in v:   return "Democrat"
    if "REPUBLICAN" in v: return "Republican"
    return "Other"


def autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)


# ── Build lookup from exploded master ─────────────────────────────────────────
print(f"Loading exploded LPAC master from {LPAC_EXPLODED}...")
df_exp = pd.read_excel(LPAC_EXPLODED, sheet_name=0, dtype=str)
df_exp.columns = [c.strip() for c in df_exp.columns]

required = {"committee_id", "committee_type"}
missing  = required - set(df_exp.columns)
if missing:
    raise ValueError(f"Exploded master missing columns: {missing}\nFound: {df_exp.columns.tolist()}")

df_exp["_join_key"] = df_exp["committee_id"].apply(clean)
df_exp["party"]     = df_exp["party"].apply(norm_party)

JOIN_COLS = {
    "committee_type": "recipient_committee_type",
    "candidate_id":   "recipient_candidate_id",
    "candidate_name": "recipient_candidate_name",
    "office":         "recipient_candidate_office",
    "state":          "recipient_candidate_state",
    "district":       "recipient_candidate_district",
    "party":          "recipient_candidate_party",
}
df_lookup = (
    df_exp[["_join_key"] + [c for c in JOIN_COLS if c in df_exp.columns]]
    .drop_duplicates(subset=["_join_key"])
    .rename(columns=JOIN_COLS)
)

lpac_keys = set(df_exp.loc[df_exp["committee_type"] == "Leadership PAC", "_join_key"])
print(f"  Committees in lookup: {len(df_lookup):,}")
print(f"  Leadership PACs:      {len(lpac_keys):,}")
print()


# ── Classify a single dataframe ───────────────────────────────────────────────
def classify_df(df_in):
    df_in["_etype"] = df_in["entity_type_desc"].apply(clean)
    df_f = df_in[df_in["_etype"].isin(KEEP_ENTITY_TYPES)].copy()
    df_f["_join_key"] = df_f["recipient_committee_id"].apply(clean)

    pac_rows = df_f[df_f["_etype"] == "POLITICAL ACTION COMMITTEE"]
    pac_ids  = set(pac_rows["_join_key"].dropna()) - {""}
    overlap  = pac_ids & lpac_keys
    print(f"  CAMPAIGN COMMITTEE rows: {(df_f['_etype'] == 'CAMPAIGN COMMITTEE').sum():,}")
    print(f"  PAC rows:                {len(pac_rows):,} | "
          f"Unique IDs: {len(pac_ids)} | LPAC matches: {len(overlap)}")

    df_f = df_f.merge(df_lookup, on="_join_key", how="left")

    mask_cc = df_f["_etype"] == "CAMPAIGN COMMITTEE"
    df_f.loc[mask_cc & df_f["recipient_committee_type"].isna(),
             "recipient_committee_type"] = "Authorized Campaign Committee"

    df_f.drop(columns=["_etype", "_join_key"], inplace=True)

    new_cols  = [c for c in df_lookup.columns if c != "_join_key"]
    base_cols = [c for c in df_f.columns if c not in new_cols]
    anchor    = "recipient_committee_id"
    if anchor in base_cols:
        i       = base_cols.index(anchor) + 1
        ordered = base_cols[:i] + new_cols + base_cols[i:]
    else:
        ordered = base_cols + new_cols

    return df_f[ordered]


# ── Build formatted summary ───────────────────────────────────────────────────
def build_summary(all_dfs):
    CYCLE_LABELS = {
        "2021": "2021-2022", "2022": "2021-2022",
        "2023": "2023-2024", "2024": "2023-2024",
        "2025": "2025-2026", "2026": "2025-2026",
    }

    rows = []
    for filename, df in all_dfs.items():
        df = df.copy()

        if "two_year_transaction_period" in df.columns:
            df["_cycle"] = pd.to_numeric(
                df["two_year_transaction_period"], errors="coerce"
            ).apply(lambda x: CYCLE_LABELS.get(str(int(x))[:4], str(int(x)))
                    if pd.notna(x) else "Unknown")
        elif "disbursement_date" in df.columns:
            df["_year"]  = pd.to_datetime(df["disbursement_date"], errors="coerce").dt.year.astype(str)
            df["_cycle"] = df["_year"].map(CYCLE_LABELS).fillna("Unknown")
        else:
            df["_cycle"] = "Unknown"

        df["disbursement_amount"] = pd.to_numeric(df["disbursement_amount"], errors="coerce")
        df["_party"] = df["recipient_candidate_party"].apply(norm_party).fillna("Unknown")

        pac_name = df["committee_name"].dropna().iloc[0] if "committee_name" in df.columns else filename
        pac_id   = df["committee_id"].dropna().iloc[0]   if "committee_id"   in df.columns else ""

        # Cycle totals: ALL rows vs MATCHED rows only
        all_cycle_totals = df.groupby("_cycle")["disbursement_amount"].sum().to_dict()

        df_matched = df[df["recipient_committee_type"].isin(
            ["Authorized Campaign Committee", "Leadership PAC"]
        )].copy()
        matched_cycle_totals = df_matched.groupby("_cycle")["disbursement_amount"].sum().to_dict()

        if df_matched.empty:
            continue

        grouped = (
            df_matched.groupby(["_cycle", "recipient_committee_type", "_party"])
            ["disbursement_amount"].sum().reset_index()
        )

        for _, r in grouped.iterrows():
            cycle      = r["_cycle"]
            comm_type  = r["recipient_committee_type"]
            party      = r["_party"]
            amount     = r["disbursement_amount"]

            all_total     = all_cycle_totals.get(cycle, 0)
            matched_total = matched_cycle_totals.get(cycle, 0)

            short_type = "Auth" if comm_type == "Authorized Campaign Committee" else "LPAC"

            rows.append({
                "PAC Name":                    pac_name,
                "PAC ID":                      pac_id,
                "Election Cycle":              cycle,
                "Segment":                     f"{short_type} — {party}",
                "Committee Type":              comm_type,
                "Party":                       party,
                "$ Amount":                    amount,
                "% of Matched Cycle Total":    (amount / matched_total) if matched_total else 0,
                "Matched Cycle Total ($)":     matched_total,
                "% of All Disbursements":      (amount / all_total) if all_total else 0,
                "All Disbursements Total ($)": all_total,
            })

    if not rows:
        return pd.DataFrame()

    df_out = pd.DataFrame(rows)
    df_out.sort_values(["PAC Name", "Election Cycle", "Committee Type", "Party"], inplace=True)
    return df_out


def write_summary_excel(summary_df, output_folder):
    from openpyxl.styles import PatternFill, Font, Alignment

    summary_path = os.path.join(output_folder, "batch_summary.xlsx")

    pct_cols    = ["% of Matched Cycle Total", "% of All Disbursements"]
    dollar_cols = ["$ Amount", "Matched Cycle Total ($)", "All Disbursements Total ($)"]

    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.sheets["Summary"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = summary_df.columns[cell.column - 1]
                if col_name in dollar_cols:
                    cell.number_format = '$#,##0'
                elif col_name in pct_cols:
                    cell.number_format = '0.0%'
                if col_name == "Segment":
                    val = str(cell.value or "")
                    if "Auth" in val and "Democrat" in val:
                        cell.font = Font(color="1F4E79", bold=True)
                    elif "Auth" in val and "Republican" in val:
                        cell.font = Font(color="C00000", bold=True)
                    elif "LPAC" in val and "Democrat" in val:
                        cell.font = Font(color="2E75B6", bold=True)
                    elif "LPAC" in val and "Republican" in val:
                        cell.font = Font(color="FF0000", bold=True)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    print(f"  Summary → {summary_path}")


# ── Process each CSV ──────────────────────────────────────────────────────────
csv_files = sorted(glob.glob(os.path.join(INPUT_FOLDER, "*.csv")))
if TEST_FILE:
    csv_files = [f for f in csv_files if os.path.basename(f) == TEST_FILE]
    if not csv_files:
        direct = os.path.join(INPUT_FOLDER, TEST_FILE)
        if os.path.exists(direct):
            csv_files = [direct]
        else:
            raise FileNotFoundError(f"Test file not found: {TEST_FILE}")
if not csv_files:
    raise FileNotFoundError(
        f"No .csv files found in: {os.path.abspath(INPUT_FOLDER)}\n"
        "Set INPUT_FOLDER to the folder containing your disbursement CSVs."
    )

print(f"Found {len(csv_files)} CSV file(s) in '{INPUT_FOLDER}'\n")
print("=" * 60)

summary_rows       = []
all_classified_dfs = {}

for i, csv_path in enumerate(csv_files, 1):
    filename = os.path.basename(csv_path)
    stem     = os.path.splitext(filename)[0]
    out_path = os.path.join(OUTPUT_FOLDER, f"{stem}_analyzed.xlsx")

    print(f"[{i}/{len(csv_files)}] {filename}")

    try:
        if zipfile.is_zipfile(csv_path):
            with zipfile.ZipFile(csv_path) as zf:
                inner = [n for n in zf.namelist() if n.endswith(".csv")]
                if not inner:
                    raise ValueError(f"No .csv inside ZIP: {csv_path}")
                with zf.open(inner[0]) as f:
                    df_raw = pd.read_csv(
                        io.TextIOWrapper(f, encoding="utf-8", errors="replace"),
                        dtype=str, low_memory=False
                    )
        else:
            df_raw = pd.read_csv(csv_path, dtype=str, low_memory=False,
                                 encoding_errors="replace")

        print(f"  Loaded {len(df_raw):,} rows")

        df_out = classify_df(df_raw)

        n_total = len(df_out)
        n_auth  = (df_out["recipient_committee_type"] == "Authorized Campaign Committee").sum()
        n_lpac  = (df_out["recipient_committee_type"] == "Leadership PAC").sum()
        n_none  = df_out["recipient_committee_type"].isna().sum()
        n_dem   = (df_out["recipient_candidate_party"] == "Democrat").sum()
        n_rep   = (df_out["recipient_candidate_party"] == "Republican").sum()
        n_unk   = df_out["recipient_candidate_party"].isna().sum()

        print(f"  Authorized:      {n_auth:,}")
        print(f"  Leadership PACs: {n_lpac:,}")
        print(f"  Unmatched PACs:  {n_none:,}")
        print(f"  Democrat:        {n_dem:,}")
        print(f"  Republican:      {n_rep:,}")
        print(f"  Party unknown:   {n_unk:,}")

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="All Disbursements")
            df_out[df_out["recipient_committee_type"].notna()].to_excel(
                writer, index=False, sheet_name="Matched Only")
            df_out[df_out["recipient_committee_type"] == "Authorized Campaign Committee"].to_excel(
                writer, index=False, sheet_name="Auth Committees")
            df_out[df_out["recipient_committee_type"] == "Leadership PAC"].to_excel(
                writer, index=False, sheet_name="Leadership PACs")
            df_out[df_out["recipient_candidate_party"] == "Democrat"].to_excel(
                writer, index=False, sheet_name="Democrat")
            df_out[df_out["recipient_candidate_party"] == "Republican"].to_excel(
                writer, index=False, sheet_name="Republican")
            unmatched = df_out[df_out["recipient_committee_type"].isna()]
            if not unmatched.empty:
                unmatched.to_excel(writer, index=False, sheet_name="Unmatched PACs")
            for ws in writer.sheets.values():
                autofit(ws)

        print(f"  Saved → {out_path}\n")
        all_classified_dfs[filename] = df_out

        summary_rows.append({
            "file":            filename,
            "total_rows":      n_total,
            "authorized":      n_auth,
            "leadership_pacs": n_lpac,
            "unmatched_pacs":  n_none,
            "democrat":        n_dem,
            "republican":      n_rep,
            "party_unknown":   n_unk,
            "status":          "OK",
        })

    except Exception as e:
        print(f"  ERROR: {e}\n")
        summary_rows.append({"file": filename, "status": f"ERROR: {e}"})


# ── Write outputs ─────────────────────────────────────────────────────────────
df_log         = pd.DataFrame(summary_rows)
batch_log_path = os.path.join(OUTPUT_FOLDER, "batch_log.xlsx")

with pd.ExcelWriter(batch_log_path, engine="openpyxl") as writer:
    df_log.to_excel(writer, index=False, sheet_name="Log")
    for ws in writer.sheets.values():
        autofit(ws)

if all_classified_dfs:
    print("Building formatted summary...")
    summary_df = build_summary(all_classified_dfs)
    if not summary_df.empty:
        write_summary_excel(summary_df, OUTPUT_FOLDER)

print("=" * 60)
print(f"Batch complete. {len(csv_files)} file(s) processed.")
print(f"Log     → {batch_log_path}")
print(f"Summary → {os.path.join(OUTPUT_FOLDER, 'batch_summary.xlsx')}")

ok = df_log[df_log["status"] == "OK"]
if not ok.empty:
    print(f"\nAggregated totals:")
    print(f"  Total filtered rows:    {ok['total_rows'].sum():,}")
    print(f"  Authorized committees:  {ok['authorized'].sum():,}")
    print(f"  Leadership PACs:        {ok['leadership_pacs'].sum():,}")
    print(f"  Unmatched PACs:         {ok['unmatched_pacs'].sum():,}")
    print(f"  Democrat:               {ok['democrat'].sum():,}")
    print(f"  Republican:             {ok['republican'].sum():,}")
    print(f"  Party unknown:          {ok['party_unknown'].sum():,}")